import pandas as pd
import re
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ──────────────────────────────────────────────
# NORMALIZACIÓN DE CLAVES
# ──────────────────────────────────────────────

def normalizar_clave(prefijo: str, folio: str) -> str:
    """
    Unifica prefijo + folio en una clave comparable sin importar el formato.
    Ejemplos:
      ("BOG", "16856")      → "BOG16856"
      ("FEMQ", "00004465")  → "FEMQ4465"
      ("", "FEMQ-00004465") → "FEMQ4465"
    """
    prefijo = str(prefijo).strip() if pd.notna(prefijo) else ""
    folio = str(folio).strip() if pd.notna(folio) else ""

    # Si el folio ya contiene el prefijo (formato Siesa: "FEMQ-00004465")
    if "-" in folio or not prefijo:
        partes = re.split(r"[-_\s]", folio, maxsplit=1)
        if len(partes) == 2:
            prefijo = partes[0]
            folio = partes[1]
        else:
            folio = partes[0]

    # Limpiar caracteres no alfanuméricos
    prefijo = re.sub(r"[^A-Za-z0-9]", "", prefijo).upper()
    # Quitar ceros a la izquierda del folio numérico
    folio_limpio = re.sub(r"[^0-9]", "", folio)
    folio_limpio = str(int(folio_limpio)) if folio_limpio else folio

    return f"{prefijo}{folio_limpio}"


def normalizar_nit(nit) -> str:
    if pd.isna(nit):
        return ""
    return re.sub(r"[^0-9]", "", str(nit)).strip()


# ──────────────────────────────────────────────
# LECTURA DE ARCHIVOS
# ──────────────────────────────────────────────

def leer_dian(contenido: bytes) -> pd.DataFrame:
    """Lee el Excel de la DIAN y retorna DataFrame normalizado."""
    extension = "xls" if contenido[:2] == b'\xd0\xcf' else "xlsx"

    df = pd.read_excel(
        io.BytesIO(contenido),
        engine="xlrd" if extension == "xls" else "openpyxl",
        dtype=str
    )

    # Validar columnas mínimas
    columnas_requeridas = ["Folio", "Prefijo", "NIT Emisor", "Nombre Emisor","Fecha Emisión"]
    for col in columnas_requeridas:
        if col not in df.columns:
            raise ValueError(f"El archivo DIAN no tiene la columna requerida: '{col}'")
    

    # Filtrar solo recibidas (columna 'Grupo' puede no existir en todos los formatos DIAN)
    if "Grupo" in df.columns:
        df = df[df["Grupo"].astype(str).str.strip().str.lower() == "recibido"]
        if df.empty:
            raise ValueError("No se encontraron facturas con Grupo = 'Recibido' en el archivo DIAN.")

    # Filtrar solo documentos que son facturas reales (excluye acuses, nóminas, etc.)
    TIPOS_FACTURA = [
        "Factura electrónica",
        "Nota de crédito electrónica",
        "Nota de débito electrónica",
        "Documento soporte con no obligados",
        "Documento equivalente POS",
        "Nota de ajuste del documento soporte",
    ]
    
    

    if "Tipo de documento" in df.columns:
        df = df[df["Tipo de documento"].isin(TIPOS_FACTURA)]

    df = df[columnas_requeridas].copy()
    df.columns = ["folio", "prefijo", "nit", "nombre","fecha"]  

    df["nit"] = df["nit"].apply(normalizar_nit)

    # Construir folio_original y clave
    def folio_original(row):
        prefijo = str(row["prefijo"]).strip() if pd.notna(row["prefijo"]) and str(row["prefijo"]).strip() not in ("", "nan") else ""
        folio = str(row["folio"]).strip() if pd.notna(row["folio"]) else ""
        if prefijo:
            return f"{prefijo}-{folio}"
        return folio

    df["folio_original"] = df.apply(folio_original, axis=1)
    df["clave"] = df.apply(lambda r: normalizar_clave(r["prefijo"], r["folio"]), axis=1)

    df = df[df["nit"] != ""].dropna(subset=["clave"])
    df = df[df["clave"] != ""]
    df = df[df["clave"] != "0"]

    return df


def leer_siesa(contenido: bytes) -> pd.DataFrame:
    """Lee el Excel de Siesa y retorna DataFrame normalizado."""
    extension = "xls"
    if contenido[:4] == b'PK\x03\x04':  # magic bytes de .xlsx
        extension = "xlsx"

    df = pd.read_excel(
        io.BytesIO(contenido),
        engine="xlrd" if extension == "xls" else "openpyxl",
        dtype=str
    )

    # Validar columnas mínimas
    columnas_requeridas = ["Proveedor", "Docto. proveedor", "Razón social proveedor"]
    for col in columnas_requeridas:
        if col not in df.columns:
            raise ValueError(f"El archivo Siesa no tiene la columna requerida: '{col}'")

    df = df[columnas_requeridas].copy()
    df.columns = ["nit", "docto", "nombre"]

    # Filtrar filas sin NIT válido (subtotales y separadores de Siesa)
    df["nit"] = df["nit"].apply(normalizar_nit)
    df = df[df["nit"].str.len() >= 6]  # NITs válidos tienen al menos 6 dígitos
    df = df[df["docto"].notna()]
    df = df[df["docto"].astype(str).str.strip() != ""]
    df = df[df["docto"].astype(str).str.strip() != "nan"]

    # Normalizar clave desde "Docto. proveedor" (ej: FEMQ-00004465)
    df["clave"] = df["docto"].apply(lambda x: normalizar_clave("", str(x)))
    df["docto_original"] = df["docto"].astype(str).str.strip()

    df = df[df["clave"] != ""]

    return df


# ──────────────────────────────────────────────
# COMPARACIÓN PRINCIPAL
# ──────────────────────────────────────────────

def _ejecutar_comparacion(df_dian: pd.DataFrame, df_siesa: pd.DataFrame) -> dict:
    """
    Motor central de comparación reutilizable.
    Acepta DataFrames ya normalizados de cualquier fuente (Siesa Excel u Odoo API).
    """
    # Match por (NIT, clave) únicamente. NO usar solo `clave` como fallback:
    # genera falsos positivos cuando dos proveedores comparten prefijo+folio.
    siesa_index = set(zip(df_siesa["nit"], df_siesa["clave"]))

    proveedores = {}
    for _, row in df_dian.iterrows():
        nit = row["nit"]
        nombre = row["nombre"]
        clave = row["clave"]
        folio_original = row["folio_original"]

        if nit not in proveedores:
            proveedores[nit] = {
                "nit": nit,
                "nombre": nombre,
                "facturas_dian": [],
                "faltantes": [],
                "encontradas": []
            }

        proveedores[nit]["facturas_dian"].append(folio_original)

        en_siesa = (nit, clave) in siesa_index

        if en_siesa:
            proveedores[nit]["encontradas"].append(folio_original)
        else:
            proveedores[nit]["faltantes"].append({
                "factura": folio_original,
                "fecha": str(row["fecha"]) if pd.notna(row["fecha"]) else "Sin fecha"
            })

    lista_proveedores = []
    for nit, datos in sorted(proveedores.items(), key=lambda x: len(x[1]["faltantes"]), reverse=True):
        lista_proveedores.append({
            "nit": nit,
            "nombre": datos["nombre"],
            "total_dian": len(datos["facturas_dian"]),
            "total_en_siesa": len(datos["encontradas"]),
            "total_faltantes": len(datos["faltantes"]),
            "faltantes": sorted(datos["faltantes"], key=lambda x: x["fecha"]),
            "encontradas": sorted(datos["encontradas"])
        })

    total_dian = sum(p["total_dian"] for p in lista_proveedores)
    total_siesa = sum(p["total_en_siesa"] for p in lista_proveedores)
    total_faltantes = sum(p["total_faltantes"] for p in lista_proveedores)

    return {
        "resumen_general": {
            "total_proveedores": len(lista_proveedores),
            "total_dian": total_dian,
            "total_en_siesa": total_siesa,
            "total_faltantes": total_faltantes,
            "porcentaje_completitud": round((total_siesa / total_dian * 100), 1) if total_dian > 0 else 0
        },
        "proveedores": lista_proveedores,
        "narrativa": ""
    }


def _odoo_to_siesa_df(facturas_odoo: list) -> pd.DataFrame:
    """
    Convierte la lista de facturas extraídas de Odoo al mismo formato
    que produce leer_siesa(), para que _ejecutar_comparacion() funcione sin cambios.
    """
    if not facturas_odoo:
        return pd.DataFrame(columns=["nit", "clave", "nombre", "docto_original"])
    rows = [
        {
            "nit": f["nit"],
            "clave": f["factura_clave"],
            "nombre": f["nombre"],
            "docto_original": f["factura_original"],
        }
        for f in facturas_odoo
        if f.get("nit") and f.get("factura_clave")
    ]
    return pd.DataFrame(rows)


def comparar_facturas(dian_bytes: bytes, siesa_bytes: bytes) -> dict:
    """Compara DIAN vs archivo Excel de Siesa."""
    df_dian = leer_dian(dian_bytes)
    df_siesa = leer_siesa(siesa_bytes)
    return _ejecutar_comparacion(df_dian, df_siesa)


def comparar_facturas_odoo(dian_bytes: bytes, facturas_odoo: list) -> dict:
    """Compara DIAN vs facturas extraídas de Odoo vía XML-RPC API."""
    df_dian = leer_dian(dian_bytes)
    df_siesa = _odoo_to_siesa_df(facturas_odoo)
    return _ejecutar_comparacion(df_dian, df_siesa)


# ──────────────────────────────────────────────
# GENERACIÓN DEL REPORTE EXCEL
# ──────────────────────────────────────────────

def generar_excel_reporte(resultado: dict) -> str:
    wb = Workbook()

    # ── Hoja 1: Resumen general ──
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"
    _estilo_resumen(ws_resumen, resultado)

    # ── Hoja 2: Detalle por proveedor ──
    ws_detalle = wb.create_sheet("Detalle por Proveedor")
    _estilo_detalle(ws_detalle, resultado)

    # ── Hoja 3: Solo faltantes ──
    ws_faltantes = wb.create_sheet("Facturas Faltantes")
    _estilo_faltantes(ws_faltantes, resultado)

    ruta = "/tmp/reporte_comparacion_facturas.xlsx"
    wb.save(ruta)
    return ruta


def _color_header(ws, fila, col_inicio, col_fin, texto, color="1A3A5C"):
    ws.merge_cells(start_row=fila, start_column=col_inicio, end_row=fila, end_column=col_fin)
    celda = ws.cell(row=fila, column=col_inicio, value=texto)
    celda.font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    celda.fill = PatternFill("solid", fgColor=color)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 22


def _borde():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _estilo_resumen(ws, resultado):
    r = resultado["resumen_general"]
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    _color_header(ws, 1, 1, 4, f"REPORTE COMPARACIÓN DIAN vs SIESA  —  {fecha}")

    ws.merge_cells("A3:D3")
    ws["A3"] = "RESUMEN EJECUTIVO"
    ws["A3"].font = Font(bold=True, size=10, name="Arial", color="1A3A5C")

    datos = [
        ("Total proveedores analizados", r["total_proveedores"], "", ""),
        ("Total facturas en DIAN", r["total_dian"], "", ""),
        ("Total encontradas en Siesa", r["total_en_siesa"], "", ""),
        ("Total faltantes en Siesa", r["total_faltantes"], "", ""),
        ("Porcentaje completitud", f"{r['porcentaje_completitud']}%", "", ""),
    ]

    for i, (label, valor, _, __) in enumerate(datos, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10)
        c = ws.cell(row=i, column=2, value=valor)
        c.font = Font(name="Arial", size=10, bold=True)
        c.alignment = Alignment(horizontal="center")
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = _borde()

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20

    # Tabla de proveedores
    fila = 11
    _color_header(ws, fila, 1, 5, "DETALLE POR PROVEEDOR", color="185FA5")
    fila += 1

    encabezados = ["NIT", "Nombre proveedor", "DIAN", "En Siesa", "Faltantes"]
    col_widths = [16, 42, 10, 10, 10]
    for col, (enc, ancho) in enumerate(zip(encabezados, col_widths), start=1):
        c = ws.cell(row=fila, column=col, value=enc)
        c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center")
        c.border = _borde()
        ws.column_dimensions[get_column_letter(col)].width = ancho

    fila += 1
    for i, p in enumerate(resultado["proveedores"]):
        color_fila = "EBF3FB" if i % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=color_fila)
        valores = [p["nit"], p["nombre"], p["total_dian"], p["total_en_siesa"], p["total_faltantes"]]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = fill
            c.border = _borde()
            if col > 2:
                c.alignment = Alignment(horizontal="center")
            if col == 5 and val > 0:
                c.font = Font(name="Arial", size=9, color="C00000", bold=True)
        fila += 1


def _estilo_detalle(ws, resultado):
    _color_header(ws, 1, 1, 6, "DETALLE COMPLETO POR PROVEEDOR")

    encabezados = ["NIT", "Nombre proveedor", "Factura DIAN", "En Siesa", "Estado"]
    col_widths = [16, 42, 18, 10, 14]

    fila = 3
    for p in resultado["proveedores"]:
        _color_header(ws, fila, 1, 5, f"{p['nombre']}  (NIT: {p['nit']})", color="2E75B6")
        fila += 1

        for col, (enc, ancho) in enumerate(zip(encabezados, col_widths), start=1):
            c = ws.cell(row=fila, column=col, value=enc)
            c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4472C4")
            c.alignment = Alignment(horizontal="center")
            c.border = _borde()
            ws.column_dimensions[get_column_letter(col)].width = ancho
        fila += 1

        todas_claves = sorted(list(set(p["encontradas"]) | set(f["factura"] for f in p["faltantes"])))
        for i, factura_clave in enumerate(todas_claves):
            encontrada = factura_clave in p["encontradas"]
            color_fila = "E2EFDA" if encontrada else "FCE4D6"
            fill = PatternFill("solid", fgColor=color_fila)
            estado_txt = "✔ Encontrada" if encontrada else "✘ Faltante"
            color_estado = "375623" if encontrada else "C00000"

            ws.cell(row=fila, column=1, value=p["nit"]).font = Font(name="Arial", size=9)
            ws.cell(row=fila, column=2, value=p["nombre"]).font = Font(name="Arial", size=9)
            ws.cell(row=fila, column=3, value=factura_clave).font = Font(name="Arial", size=9)
            ws.cell(row=fila, column=4, value="Sí" if encontrada else "No").alignment = Alignment(horizontal="center")
            c_estado = ws.cell(row=fila, column=5, value=estado_txt)
            c_estado.font = Font(name="Arial", size=9, bold=True, color=color_estado)
            c_estado.alignment = Alignment(horizontal="center")

            for col in range(1, 6):
                ws.cell(row=fila, column=col).fill = fill
                ws.cell(row=fila, column=col).border = _borde()
            fila += 1

        fila += 1  # Espacio entre proveedores


def _estilo_faltantes(ws, resultado):
    _color_header(ws, 1, 1, 4, "FACTURAS FALTANTES EN SIESA", color="C00000")

    encabezados = ["NIT", "Nombre proveedor", "Factura faltante", "Prefijo-Folio"]
    col_widths = [16, 42, 20, 16]

    fila = 3
    for col, (enc, ancho) in enumerate(zip(encabezados, col_widths), start=1):
        c = ws.cell(row=fila, column=col, value=enc)
        c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C00000")
        c.alignment = Alignment(horizontal="center")
        c.border = _borde()
        ws.column_dimensions[get_column_letter(col)].width = ancho
    fila += 1

    i = 0
    for p in resultado["proveedores"]:
        for f_obj in p["faltantes"]:
            color_fila = "FCE4D6" if i % 2 == 0 else "FFFFFF"
            fill = PatternFill("solid", fgColor=color_fila)
            valores = [p["nit"], p["nombre"], f_obj["factura"], f_obj["factura"]]
            for col, val in enumerate(valores, start=1):
                c = ws.cell(row=fila, column=col, value=val)
                c.font = Font(name="Arial", size=9)
                c.fill = fill
                c.border = _borde()
            fila += 1
            i += 1

    if fila == 4:
        ws.merge_cells("A4:D4")
        c = ws.cell(row=4, column=1, value="✅ No hay facturas faltantes. ¡Todos los registros están completos!")
        c.font = Font(name="Arial", size=10, bold=True, color="375623")
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        c.alignment = Alignment(horizontal="center")
